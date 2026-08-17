"""Controlled four-model benchmark for LLM material normalization and process selection.

This experiment is intentionally separate from the FastAPI application.  It
uses the same 35 BOM entries and the exported 608-process ELCD/openLCA catalog,
but every model receives identical deterministic candidate lists and an
identical prompt.  No emission-factor prediction is performed here.
"""

from __future__ import annotations

import argparse
import csv
import gc
import json
import os
import random
import time
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from prompts import SYSTEM_PROMPT, build_prompt
from retrieval import CatalogProcess, retrieve_candidates
from utils import canonical_text, environment_metadata, extract_json_object, utc_now_iso, write_json


HERE = Path(__file__).resolve().parent
REPO_ROOT = HERE.parents[1]
DEFAULT_BOM = HERE / "bom_35_items.csv"
DEFAULT_CATALOG = REPO_ROOT / "research_artifacts" / "openlca" / "ELCD_Process_Catalog.xlsx"
DEFAULT_REGISTRY = HERE / "model_registry.json"
DEFAULT_RESULTS = HERE / "results"

ALLOWED_DECISIONS = {"Direct", "Proxy", "Review Required"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--model", required=True, choices=["llama", "qwen", "deepseek", "mistral"])
    parser.add_argument("--repeats", type=int, default=5)
    parser.add_argument("--top-k", type=int, default=10)
    parser.add_argument("--seed", type=int, default=20260816)
    parser.add_argument("--max-new-tokens", type=int, default=192)
    parser.add_argument("--load-in-4bit", action="store_true", help="Recommended for Colab T4.")
    parser.add_argument("--bom", type=Path, default=DEFAULT_BOM)
    parser.add_argument("--catalog", type=Path, default=DEFAULT_CATALOG)
    parser.add_argument("--registry", type=Path, default=DEFAULT_REGISTRY)
    parser.add_argument("--output-root", type=Path, default=DEFAULT_RESULTS)
    parser.add_argument("--run-name", type=str, default=None)
    parser.add_argument("--resume", action="store_true")
    parser.add_argument("--limit-items", type=int, default=None, help="Smoke-test only; omit for final benchmark.")
    return parser.parse_args()


def load_registry(path: Path) -> dict[str, dict[str, Any]]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def load_bom(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    required = {"id", "case_study", "original_description", "quantity", "unit"}
    if not rows or not required.issubset(rows[0]):
        raise ValueError(f"BOM CSV must contain columns: {sorted(required)}")
    return rows


def load_catalog(path: Path) -> list[CatalogProcess]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["ELCD Processes"] if "ELCD Processes" in workbook.sheetnames else workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    index = {name: i for i, name in enumerate(headers)}
    required = {"Process UUID", "Process Name"}
    missing = required - index.keys()
    if missing:
        raise ValueError(f"Catalog is missing required columns: {sorted(missing)}")

    catalog: list[CatalogProcess] = []
    for row in rows:
        uuid = str(row[index["Process UUID"]] or "").strip()
        name = str(row[index["Process Name"]] or "").strip()
        if not uuid or not name:
            continue
        def optional(column: str) -> str:
            i = index.get(column)
            return str(row[i] or "").strip() if i is not None else ""
        catalog.append(
            CatalogProcess(
                uuid=uuid,
                name=name,
                location=optional("Location"),
                process_type=optional("Process Type"),
                category=optional("Category"),
            )
        )
    workbook.close()
    if not catalog:
        raise ValueError("No process records were loaded from the catalog")
    return catalog


def resolve_hf_revision(model_id: str, requested_revision: str | None) -> str | None:
    try:
        from huggingface_hub import model_info
        info = model_info(model_id, revision=requested_revision)
        return getattr(info, "sha", None)
    except Exception:
        return requested_revision


def load_model(model_id: str, revision: str | None, load_in_4bit: bool):
    import torch
    import transformers
    from transformers import AutoModelForCausalLM, AutoTokenizer

    if load_in_4bit and not torch.cuda.is_available():
        raise RuntimeError("--load-in-4bit requires a CUDA GPU")

    tokenizer = AutoTokenizer.from_pretrained(model_id, revision=revision)
    model_kwargs: dict[str, Any] = {
        "revision": revision,
        "device_map": "auto",
        "low_cpu_mem_usage": True,
    }
    if load_in_4bit:
        from transformers import BitsAndBytesConfig
        model_kwargs["quantization_config"] = BitsAndBytesConfig(
            load_in_4bit=True,
            bnb_4bit_quant_type="nf4",
            bnb_4bit_use_double_quant=True,
            bnb_4bit_compute_dtype=torch.float16,
        )
    elif torch.cuda.is_available():
        model_kwargs["torch_dtype"] = torch.float16

    model = AutoModelForCausalLM.from_pretrained(model_id, **model_kwargs)
    model.eval()

    if tokenizer.pad_token_id is None:
        tokenizer.pad_token_id = tokenizer.eos_token_id
    return torch, transformers, tokenizer, model


def prepare_input(tokenizer, system_prompt: str, user_prompt: str, device):
    # The benchmark instructions are embedded in user_prompt so every model sees
    # the same instruction text.  We intentionally use a single user message
    # because some compared chat models (notably deepseek-llm-7b-chat) do not
    # recommend a system-role message.  Only the model-specific chat wrapper
    # differs, as required by each tokenizer.
    messages = [{"role": "user", "content": user_prompt}]
    chat_template = getattr(tokenizer, "chat_template", None)
    if chat_template:
        text = tokenizer.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
    else:
        text = f"User:\n{user_prompt}\n\nAssistant:\n"
    encoded = tokenizer(text, return_tensors="pt")
    return {key: value.to(device) for key, value in encoded.items()}


def generate_once(torch, tokenizer, model, prompt: str, max_new_tokens: int) -> tuple[str, float, int, int]:
    device = next(model.parameters()).device
    inputs = prepare_input(tokenizer, SYSTEM_PROMPT, prompt, device)
    prompt_length = inputs["input_ids"].shape[-1]
    if torch.cuda.is_available():
        torch.cuda.synchronize()
    start = time.perf_counter()
    with torch.inference_mode():
        generated = model.generate(
            **inputs,
            max_new_tokens=max_new_tokens,
            do_sample=False,
            pad_token_id=tokenizer.pad_token_id,
            eos_token_id=tokenizer.eos_token_id,
        )
    if torch.cuda.is_available():
        torch.cuda.synchronize()
    elapsed = time.perf_counter() - start
    output_token_count = int(generated[0].shape[-1] - prompt_length)
    text = tokenizer.decode(generated[0][prompt_length:], skip_special_tokens=True).strip()
    return text, elapsed, int(prompt_length), output_token_count


def validate_output(data: dict[str, Any], candidates: list[CatalogProcess]) -> dict[str, Any]:
    normalized = data.get("normalized_material")
    if not isinstance(normalized, str) or not normalized.strip():
        raise ValueError("normalized_material is missing or empty")

    decision = data.get("decision")
    if not isinstance(decision, str):
        raise ValueError("decision is missing")
    normalized_decision = decision.strip().lower()
    decision_map = {
        "direct": "Direct",
        "proxy": "Proxy",
        "review required": "Review Required",
        "review_required": "Review Required",
        "review-required": "Review Required",
    }
    decision = decision_map.get(normalized_decision, decision.strip())
    if decision not in ALLOWED_DECISIONS:
        raise ValueError(f"invalid decision: {decision!r}")

    try:
        selected_index = int(data.get("selected_candidate"))
    except (TypeError, ValueError) as exc:
        raise ValueError("selected_candidate is not an integer") from exc

    # Critical benchmark behavior: -1 remains Review Required.  It is never
    # silently converted to candidate 0.
    if selected_index == -1:
        if decision != "Review Required":
            raise ValueError("selected_candidate=-1 requires decision='Review Required'")
        selected = None
    elif 0 <= selected_index < len(candidates):
        if decision == "Review Required":
            raise ValueError("Review Required must use selected_candidate=-1")
        selected = candidates[selected_index]
    else:
        raise ValueError(f"selected_candidate {selected_index} is outside candidate range")

    reason = data.get("reason")
    return {
        "normalized_material": normalized.strip(),
        "selected_candidate": selected_index,
        "decision": decision,
        "reason": reason.strip() if isinstance(reason, str) else "",
        "selected_process_uuid": selected.uuid if selected else "",
        "selected_process_name": selected.name if selected else "",
        "selected_process_location": selected.location if selected else "",
    }


def existing_keys(path: Path) -> set[tuple[str, int]]:
    if not path.exists():
        return set()
    keys: set[tuple[str, int]] = set()
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            if not line.strip():
                continue
            try:
                row = json.loads(line)
                keys.add((str(row["bom_id"]), int(row["repeat"])))
            except Exception:
                continue
    return keys


def jsonl_to_csv(jsonl_path: Path, csv_path: Path) -> list[dict[str, Any]]:
    rows = []
    with jsonl_path.open("r", encoding="utf-8") as handle:
        for line in handle:
            if line.strip():
                rows.append(json.loads(line))
    if not rows:
        return rows
    columns = list(rows[0].keys())
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    return rows


def write_repeatability(rows: list[dict[str, Any]], path: Path) -> None:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[str(row["bom_id"])].append(row)

    output = []
    for bom_id, items in sorted(grouped.items()):
        valid = [row for row in items if row.get("parse_ok")]
        signatures = {
            (
                canonical_text(str(row.get("normalized_material") or "")),
                str(row.get("selected_process_uuid") or ""),
                str(row.get("decision") or ""),
            )
            for row in valid
        }
        output.append(
            {
                "bom_id": bom_id,
                "runs_requested": len(items),
                "valid_runs": len(valid),
                "unique_valid_outputs": len(signatures),
                "all_valid_runs_identical": bool(valid) and len(signatures) == 1 and len(valid) == len(items),
            }
        )
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(output[0].keys()) if output else ["bom_id"])
        writer.writeheader()
        writer.writerows(output)


def main() -> None:
    args = parse_args()
    if args.repeats < 1:
        raise SystemExit("--repeats must be >= 1")
    if args.top_k < 1:
        raise SystemExit("--top-k must be >= 1")

    registry = load_registry(args.registry)
    spec = registry[args.model]
    model_id = spec["model_id"]
    requested_revision = spec.get("revision")
    resolved_revision = resolve_hf_revision(model_id, requested_revision)

    bom = load_bom(args.bom)
    if args.limit_items is not None:
        bom = bom[: args.limit_items]
    catalog = load_catalog(args.catalog)

    run_name = args.run_name or f"{args.model}_{time.strftime('%Y%m%d_%H%M%S')}"
    run_dir = args.output_root / run_name
    run_dir.mkdir(parents=True, exist_ok=True)
    raw_jsonl = run_dir / "raw_results.jsonl"

    completed = existing_keys(raw_jsonl) if args.resume else set()
    if raw_jsonl.exists() and not args.resume and raw_jsonl.stat().st_size > 0:
        raise SystemExit(f"{raw_jsonl} already contains results. Use --resume or a new --run-name.")

    # Prepare identical deterministic candidate lists before loading the model.
    candidate_map: dict[str, list[CatalogProcess]] = {}
    candidate_rows = []
    for item in bom:
        candidates = retrieve_candidates(item["original_description"], catalog, limit=args.top_k)
        candidate_map[item["id"]] = candidates
        candidate_rows.append(
            {
                "bom_id": item["id"],
                "original_description": item["original_description"],
                "top_k": args.top_k,
                "candidate_uuids_json": json.dumps([p.uuid for p in candidates], ensure_ascii=False),
                "candidate_names_json": json.dumps([p.name for p in candidates], ensure_ascii=False),
                "candidate_locations_json": json.dumps([p.location for p in candidates], ensure_ascii=False),
            }
        )
    with (run_dir / "candidate_retrieval.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(candidate_rows[0].keys()))
        writer.writeheader()
        writer.writerows(candidate_rows)

    print(f"Model: {spec['display_name']} ({model_id})")
    print(f"BOM items: {len(bom)} | repeats: {args.repeats} | top-k: {args.top_k}")
    print(f"Catalog processes: {len(catalog)}")
    print(f"Loading model... 4-bit={args.load_in_4bit}")

    torch, transformers, tokenizer, model = load_model(model_id, resolved_revision, args.load_in_4bit)
    actual_revision = getattr(getattr(model, "config", None), "_commit_hash", None) or resolved_revision

    random.seed(args.seed)
    torch.manual_seed(args.seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(args.seed)
        torch.cuda.reset_peak_memory_stats()

    manifest = {
        "benchmark": "LLM material normalization and ELCD process-selection benchmark",
        "model_alias": args.model,
        "display_name": spec["display_name"],
        "model_id": model_id,
        "requested_revision": requested_revision,
        "resolved_revision_sha": actual_revision,
        "parameter_class": spec.get("parameter_class"),
        "model_notes": spec.get("notes"),
        "bom_file": str(args.bom.relative_to(REPO_ROOT) if args.bom.is_relative_to(REPO_ROOT) else args.bom),
        "catalog_file": str(args.catalog.relative_to(REPO_ROOT) if args.catalog.is_relative_to(REPO_ROOT) else args.catalog),
        "catalog_process_count": len(catalog),
        "bom_item_count": len(bom),
        "repeats": args.repeats,
        "top_k": args.top_k,
        "seed": args.seed,
        "max_new_tokens": args.max_new_tokens,
        "decoding": {"do_sample": False, "temperature": 0.0},
        "quantization": "4-bit NF4 double-quantization, fp16 compute" if args.load_in_4bit else "none",
        "batch_size": 1,
        "model_parameter_count": int(sum(parameter.numel() for parameter in model.parameters())),
        "first_parameter_dtype": str(next(model.parameters()).dtype),
        "benchmark_instruction_text": SYSTEM_PROMPT,
        "chat_role_policy": "single user message for all models; model-specific tokenizer chat wrapper only",
        "environment": environment_metadata(torch, transformers),
        "run_started_utc": utc_now_iso(),
    }
    write_json(run_dir / "run_manifest.json", manifest)

    total = len(bom) * args.repeats
    done = 0
    with raw_jsonl.open("a", encoding="utf-8") as handle:
        for repeat in range(1, args.repeats + 1):
            for item in bom:
                key = (item["id"], repeat)
                done += 1
                if key in completed:
                    print(f"[{done}/{total}] skip {item['id']} repeat {repeat} (already completed)")
                    continue

                candidates = candidate_map[item["id"]]
                prompt = build_prompt(
                    bom_id=item["id"],
                    description=item["original_description"],
                    unit=item["unit"],
                    candidates=[
                        {
                            "name": p.name,
                            "location": p.location,
                            "process_type": p.process_type,
                        }
                        for p in candidates
                    ],
                )

                row: dict[str, Any] = {
                    "model_alias": args.model,
                    "model_id": model_id,
                    "model_revision_sha": actual_revision or "",
                    "bom_id": item["id"],
                    "case_study": item["case_study"],
                    "original_description": item["original_description"],
                    "quantity": item["quantity"],
                    "unit": item["unit"],
                    "repeat": repeat,
                    "top_k": args.top_k,
                    "candidate_uuids_json": json.dumps([p.uuid for p in candidates], ensure_ascii=False),
                    "candidate_names_json": json.dumps([p.name for p in candidates], ensure_ascii=False),
                    "prompt_text": prompt,
                    "normalized_material": "",
                    "selected_candidate": "",
                    "selected_process_uuid": "",
                    "selected_process_name": "",
                    "selected_process_location": "",
                    "decision": "",
                    "reason": "",
                    "parse_ok": False,
                    "inference_seconds": "",
                    "input_token_count": "",
                    "output_token_count": "",
                    "error": "",
                    "raw_response": "",
                }

                try:
                    raw, elapsed, input_tokens, output_tokens = generate_once(torch, tokenizer, model, prompt, args.max_new_tokens)
                    row["raw_response"] = raw
                    row["inference_seconds"] = round(elapsed, 6)
                    row["input_token_count"] = input_tokens
                    row["output_token_count"] = output_tokens
                    parsed = extract_json_object(raw)
                    validated = validate_output(parsed, candidates)
                    row.update(validated)
                    row["parse_ok"] = True
                except Exception as exc:
                    row["error"] = f"{type(exc).__name__}: {exc}"
                    if torch.cuda.is_available():
                        torch.cuda.empty_cache()

                handle.write(json.dumps(row, ensure_ascii=False) + "\n")
                handle.flush()
                status = "OK" if row["parse_ok"] else "ERROR"
                print(f"[{done}/{total}] {item['id']} repeat {repeat}: {status} ({row['inference_seconds']} s)")

    rows = jsonl_to_csv(raw_jsonl, run_dir / "raw_results.csv")
    write_repeatability(rows, run_dir / "repeatability_summary.csv")

    valid = sum(bool(row.get("parse_ok")) for row in rows)
    identical_rows = []
    with (run_dir / "repeatability_summary.csv").open("r", encoding="utf-8-sig") as handle:
        identical_rows = list(csv.DictReader(handle))
    identical_items = sum(str(row.get("all_valid_runs_identical", "")).lower() == "true" for row in identical_rows)

    manifest["run_completed_utc"] = utc_now_iso()
    manifest["valid_output_count"] = valid
    manifest["total_output_count"] = len(rows)
    manifest["parse_success_rate"] = (valid / len(rows)) if rows else 0.0
    manifest["items_with_identical_outputs_across_all_repeats"] = identical_items
    manifest["item_repeatability_rate"] = (identical_items / len(identical_rows)) if identical_rows else 0.0
    if torch.cuda.is_available():
        manifest["peak_gpu_memory_gib"] = round(torch.cuda.max_memory_allocated() / (1024**3), 3)
    write_json(run_dir / "run_manifest.json", manifest)

    del model
    gc.collect()
    if torch.cuda.is_available():
        torch.cuda.empty_cache()

    print("\nBenchmark complete.")
    print(f"Results: {run_dir}")
    print(f"Parse success: {valid}/{len(rows)}")
    print(f"Items identical across all repeats: {identical_items}/{len(identical_rows)}")


if __name__ == "__main__":
    main()
