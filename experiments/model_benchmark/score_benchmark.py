"""Score completed model-benchmark outputs against the reconciled expert reference.

Run this only after Expert A and Expert B have been reconciled.  The script does
not invent missing labels.  If the final reference is incomplete, it exits and
reports how many rows still need adjudication.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook

from utils import canonical_text


HERE = Path(__file__).resolve().parent
REPO_ROOT = HERE.parents[1]
DEFAULT_REFERENCE = REPO_ROOT / "research_artifacts" / "expert_reference" / "LLM_LCA_Expert_Reference_Set_With_ELCD.xlsx"
DEFAULT_RESULTS = HERE / "results"


def parse_args():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--reference", type=Path, default=DEFAULT_REFERENCE)
    parser.add_argument("--results-root", type=Path, default=DEFAULT_RESULTS)
    parser.add_argument("--output-dir", type=Path, default=HERE / "scored_results")
    return parser.parse_args()


def token_f1(a: str, b: str) -> float:
    left = set(re.findall(r"[a-z0-9]+", (a or "").lower()))
    right = set(re.findall(r"[a-z0-9]+", (b or "").lower()))
    if not left and not right:
        return 1.0
    if not left or not right:
        return 0.0
    overlap = len(left & right)
    precision = overlap / len(left)
    recall = overlap / len(right)
    return 2 * precision * recall / (precision + recall) if precision + recall else 0.0


def load_reference(path: Path) -> dict[str, dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=False)
    if "Reconciliation" not in wb.sheetnames or "ELCD_Catalog" not in wb.sheetnames:
        raise ValueError("Reference workbook must contain Reconciliation and ELCD_Catalog sheets")

    catalog_sheet = wb["ELCD_Catalog"]
    label_to_uuid: dict[str, str] = {}
    for row in catalog_sheet.iter_rows(min_row=2, values_only=True):
        label = str(row[0] or "").strip()
        uuid = str(row[1] or "").strip()
        if label and uuid:
            label_to_uuid[label] = uuid

    ref_sheet = wb["Reconciliation"]
    reference: dict[str, dict[str, str]] = {}
    for row in ref_sheet.iter_rows(min_row=7, max_row=41, values_only=True):
        bom_id = str(row[0] or "").strip()
        if not bom_id:
            continue
        final_normalized = str(row[9] or "").strip()   # J
        final_process_label = str(row[10] or "").strip()  # K
        final_decision = str(row[12] or "").strip()   # M
        final_uuid = label_to_uuid.get(final_process_label, "")
        reference[bom_id] = {
            "reference_normalized": final_normalized,
            "reference_process_label": final_process_label,
            "reference_process_uuid": final_uuid,
            "reference_decision": final_decision,
        }
    wb.close()
    return reference


def load_results(root: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for path in sorted(root.glob("*/raw_results.csv")):
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows.extend(csv.DictReader(handle))
    return rows


def write_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        return
    columns = list(rows[0].keys())
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    args = parse_args()
    reference = load_reference(args.reference)
    incomplete = [
        bom_id
        for bom_id, row in reference.items()
        if not row["reference_normalized"] or not row["reference_decision"]
    ]
    if incomplete:
        raise SystemExit(
            f"Expert reference is not ready: {len(reference) - len(incomplete)}/{len(reference)} rows have final labels. "
            f"Still incomplete: {', '.join(incomplete)}"
        )

    results = load_results(args.results_root)
    if not results:
        raise SystemExit(f"No raw_results.csv files found under {args.results_root}")

    item_scores = []
    for row in results:
        ref = reference.get(row["bom_id"])
        if ref is None:
            continue
        try:
            candidates = json.loads(row.get("candidate_uuids_json") or "[]")
        except json.JSONDecodeError:
            candidates = []
        ref_uuid = ref["reference_process_uuid"]
        ref_decision = ref["reference_decision"]
        model_decision = row.get("decision") or ""
        selected_uuid = row.get("selected_process_uuid") or ""

        reference_requires_process = ref_decision != "Review Required" and bool(ref_uuid)
        retrieval_hit = bool(reference_requires_process and ref_uuid in candidates)
        process_exact = bool(reference_requires_process and selected_uuid == ref_uuid)
        review_required_correct = bool(
            (ref_decision == "Review Required" and model_decision == "Review Required")
            or (ref_decision != "Review Required" and model_decision != "Review Required")
        )
        decision_exact = model_decision == ref_decision
        norm_exact = canonical_text(row.get("normalized_material") or "") == canonical_text(ref["reference_normalized"])

        scored = dict(row)
        scored.update(ref)
        scored.update(
            {
                "normalization_canonical_exact": norm_exact,
                "normalization_token_f1": round(token_f1(row.get("normalized_material") or "", ref["reference_normalized"]), 6),
                "candidate_retrieval_hit": retrieval_hit,
                "process_selection_exact": process_exact,
                "decision_exact": decision_exact,
                "review_required_binary_correct": review_required_correct,
                "process_selection_eligible": reference_requires_process,
                "conditional_selection_eligible": retrieval_hit,
                "conditional_selection_exact": bool(retrieval_hit and process_exact),
            }
        )
        item_scores.append(scored)

    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in item_scores:
        grouped[row["model_alias"]].append(row)

    summaries = []
    for model_alias, rows in sorted(grouped.items()):
        valid = [r for r in rows if str(r.get("parse_ok", "")).lower() == "true"]
        process_eligible = [r for r in valid if r["process_selection_eligible"]]
        conditional = [r for r in valid if r["conditional_selection_eligible"]]
        first_repeat = [r for r in valid if str(r.get("repeat")) == "1"]
        # Candidate retrieval is deterministic; report it once per BOM (repeat 1).
        first_process_eligible = [r for r in first_repeat if r["process_selection_eligible"]]

        summaries.append(
            {
                "model_alias": model_alias,
                "valid_outputs": len(valid),
                "normalization_canonical_exact_accuracy": round(mean(bool(r["normalization_canonical_exact"]) for r in valid), 6) if valid else "",
                "normalization_mean_token_f1": round(mean(float(r["normalization_token_f1"]) for r in valid), 6) if valid else "",
                "candidate_retrieval_topk_recall": round(mean(bool(r["candidate_retrieval_hit"]) for r in first_process_eligible), 6) if first_process_eligible else "",
                "process_selection_end_to_end_accuracy": round(mean(bool(r["process_selection_exact"]) for r in process_eligible), 6) if process_eligible else "",
                "process_selection_conditional_accuracy": round(mean(bool(r["conditional_selection_exact"]) for r in conditional), 6) if conditional else "",
                "decision_accuracy": round(mean(bool(r["decision_exact"]) for r in valid), 6) if valid else "",
                "review_required_binary_accuracy": round(mean(bool(r["review_required_binary_correct"]) for r in valid), 6) if valid else "",
                "mean_inference_seconds": round(mean(float(r["inference_seconds"]) for r in valid if r.get("inference_seconds")), 6) if valid else "",
            }
        )

    args.output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(args.output_dir / "benchmark_item_scores.csv", item_scores)
    write_csv(args.output_dir / "benchmark_model_summary.csv", summaries)
    print(f"Wrote item scores: {args.output_dir / 'benchmark_item_scores.csv'}")
    print(f"Wrote model summary: {args.output_dir / 'benchmark_model_summary.csv'}")
    print("Note: canonical exact-match and token-F1 are automated normalization metrics. Human semantic adjudication can be added after the expert reference is frozen.")


if __name__ == "__main__":
    main()
